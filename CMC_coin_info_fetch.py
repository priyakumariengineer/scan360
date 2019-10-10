#Author : Parth Pandya
#Name : CMC_Coin_Info_Fetch.py
#Description : 
#To pull out twitter handle of all the currency and make a xls file. This uses APIs provided by CMC.
#This twitter handles will be used for monitoring using twitter APIs in other module.
#
#
#production link: https://pro-api.coinmarketcap.com
#Test link: https://sandbox-api.coinmarketcap.com/
#
#Documentation : https://coinmarketcap.com/api/documentation/v1/
#Pro usage Tracking : https://pro.coinmarketcap.com/account
#Sandbox Tracking : https://sandbox.coinmarketcap.com/account
#
#This example uses Python 2.7 and the python-request library.

from requests import Request, Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
import json
import csv
import math
import openpyxl 

#sandbox
url_map = 'https://sandbox-api.coinmarketcap.com/v1/cryptocurrency/map'
url_info = 'https://sandbox-api.coinmarketcap.com/v1/cryptocurrency/info'

#production
#url_map  = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/map'
#url_info = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/info'

parameters_map = {
}

parameters_info_ids = {
  'id': '1,2'
}

headers = {
  'Accepts': 'application/json',
  #TEST
  'X-CMC_PRO_API_KEY': 'Your key here',
  #PRO
  #'X-CMC_PRO_API_KEY': 'Your key here',
}


Map_details_list = [  'Dummy',
                      'Main_id' , 
                      'Name', 
                      'Symbol', 
                      'URL_name', 
                      'Is_Active', 
                      'status',
                      'first_historical_data', 
                      'last_historical_data', 
                      'platform', 
                      'platform_name', 
                      'platform_token_address' , 
                      'catagory_coin_tocken', 
                      'logo_url',
                      'description',
                      'date_added',
                      'notice',
                      'tags_minable_or_not',
                      'website',
                      'technical_doc',
                      'explorer',
                      'source_code',
                      'message_board',
                      'chat',
                      'announcement',
                      'reddit',
                      'twitter'
                    ]
					
					
session = Session()
session.headers.update(headers)

#creating an excel file for writing data.
wb = openpyxl.Workbook()
sheet = wb.active 
sheet.title = "CMC_details"

# writing titles in excel file
# carefull about index
for i in range(1, len(Map_details_list)):
  c1 = sheet.cell(row = 1, column = i)
  c1.value = Map_details_list[i]  

#fetching list of availible currency and ids.

try:
  response = session.get(url_map, params=parameters_map)
  data_map = json.loads(response.text)
  data_map_list = data_map['data']
  length_map = len(data_map_list)
  id_list = []
  
  for i in range(length_map): 
    id_list.append(data_map_list[i]['id'])
	
    sheet.cell(row = i+2, column = Map_details_list.index('Main_id')).value = data_map_list[i]['id']
    sheet.cell(row = i+2, column = Map_details_list.index('Name')).value = data_map_list[i]['name']
    sheet.cell(row = i+2, column = Map_details_list.index('Symbol')).value = data_map_list[i]['symbol']
    sheet.cell(row = i+2, column = Map_details_list.index('URL_name')).value = data_map_list[i]['slug']
    sheet.cell(row = i+2, column = Map_details_list.index('Is_Active')).value = data_map_list[i]['is_active']
    sheet.cell(row = i+2, column = Map_details_list.index('status')).value = "NA"
    sheet.cell(row = i+2, column = Map_details_list.index('first_historical_data')).value = \
    data_map_list[i]['first_historical_data']
    sheet.cell(row = i+2, column = Map_details_list.index('last_historical_data')).value = \
    data_map_list[i]['last_historical_data']
	
    if data_map_list[i]['platform']:
      curr_platform_dict = data_map_list[i]['platform']
      
      sheet.cell(row = i+2, column = Map_details_list.index('platform')).value = "YES"
      sheet.cell(row = i+2, column = Map_details_list.index('platform_name')).value = \
      curr_platform_dict['name']
      sheet.cell(row = i+2, column = Map_details_list.index('platform_token_address')).value = \
      curr_platform_dict['token_address']
      
    else:
    
      sheet.cell(row = i+2, column = Map_details_list.index('platform')).value = "NA"
      
	
  #print(id_list)
  #print(data['data'][1]['id'])
  #print(data['data'])
  
  
  
except (ConnectionError, Timeout, TooManyRedirects) as e:
  print("There is some issue fetching ID map , here is the error !!! " )
  print(e)
  
#fetching list of twitter accounts

length_id = len(id_list)

for i in range(0, length_id, 100 ):

  Curr_id_list = list(id_list[i: i+100 if i+100 < length_id else length_id ])
  parameters_info_ids['id'] = ','.join(map(str, Curr_id_list)) 

  try:
    response = session.get(url_info, params=parameters_info_ids)
    data_info = json.loads(response.text)
	
    cnt = 0
    for j in range(i, i+100 if i+100 < length_id else length_id ):

      sheet.cell(row = j+2, column = Map_details_list.index('catagory_coin_tocken')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['category'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('logo_url')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['logo'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('description')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['description'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('date_added')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['date_added'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('notice')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['notice'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('tags_minable_or_not')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['tags'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('website')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['website'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('technical_doc')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['technical_doc'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('explorer')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['explorer'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('source_code')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['source_code'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('message_board')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['message_board'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('chat')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['chat'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('announcement')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['announcement'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('reddit')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['reddit'])
      
      sheet.cell(row = j+2, column = Map_details_list.index('twitter')).value = \
      str(data_info['data'][str(Curr_id_list[cnt])]['urls']['twitter'])

      cnt = cnt + 1
	
    #data_info = str(data_info).encode('utf8')
    #print(data_info)
    #print("************************************************************")
    #print("************************************************************")
	
  except (ConnectionError, Timeout, TooManyRedirects) as e:
    print("There is some issue fetching meta data of currency , here is the error !!! " )
    print(e)
	
wb.save("CMC_Coin_Details_Auto.xlsx") 
	
